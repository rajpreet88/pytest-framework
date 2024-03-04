import openpyxl


class HomePageTestData:
    test_data_list = []

    @staticmethod
    def get_excel_data():
        book = openpyxl.load_workbook("./HomePageTestData.xlsx")
        sheet = book.active

        for i in range(2, sheet.max_row + 1):
            dict = {}
            for j in range(2, sheet.max_column + 1):
                dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            HomePageTestData.test_data_list.append(dict)
        return HomePageTestData.test_data_list

# print(HomePageTestData.get_excel_data())
